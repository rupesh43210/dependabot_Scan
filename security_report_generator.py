#!/usr/bin/env python3
"""
Security Report Generator - Enhanced Version

Professional security report generator with advanced analytics, trend analysis,
and interactive visualizations for comprehensive vulnerability management.

Features:
- Executive summary with KPIs and trends
- Detailed technical reports with resolution tracking
- Advanced analytics and predictive insights
- Interactive Excel dashboards with charts
- Compliance reporting and risk scoring
- Vulnerability lifecycle analysis

Author: GitHub Copilot
Version: 2.1.0
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import json
import warnings
warnings.filterwarnings('ignore')

# Import enhanced features
try:
    from enhanced_report_features import ComplianceReporter, TrendAnalyzer, RiskAnalyzer, SecurityMetrics
    ENHANCED_FEATURES_AVAILABLE = True
except ImportError:
    ENHANCED_FEATURES_AVAILABLE = False
    print("ℹ️  Enhanced features not available. Basic reporting mode enabled.")


class SecurityReportGenerator:
    """
    Enhanced security report generator with advanced analytics and visualizations.
    
    Features:
    - Executive summary reports with KPIs
    - Detailed technical reports with lifecycle tracking
    - Advanced analytics and trend analysis
    - Professional Excel formatting with charts
    - Risk scoring and prioritization
    - Compliance reporting capabilities
    - Vulnerability aging and resolution metrics
    """
    
    # Enhanced risk scoring weights
    SEVERITY_WEIGHTS = {
        'CRITICAL': 50,
        'HIGH': 20,
        'MEDIUM': 5,
        'LOW': 1
    }
    
    # Repository risk multipliers
    REPO_RISK_FACTORS = {
        'public': 2.0,
        'production': 1.5,
        'archived': 0.5,
        'private': 1.0
    }
    
    # Excel color schemes for enhanced formatting
    SEVERITY_COLORS = {
        'CRITICAL': {'bg': 'FF0000', 'font': 'FFFFFF'},  # Red
        'HIGH': {'bg': 'FF8C00', 'font': 'FFFFFF'},      # Orange
        'MEDIUM': {'bg': 'FFD700', 'font': '000000'},    # Yellow
        'LOW': {'bg': '90EE90', 'font': '000000'}        # Light Green
    }
    
    # Status colors for resolution tracking
    STATUS_COLORS = {
        'open': {'bg': 'FFB6C1', 'font': '000000'},      # Light Pink
        'fixed': {'bg': '98FB98', 'font': '000000'},     # Pale Green
        'dismissed': {'bg': 'D3D3D3', 'font': '000000'} # Light Gray
    }
    
    def __init__(self, scoped_repositories: Optional[List[str]] = None):
        """
        Initialize the enhanced report generator.
        
        Args:
            scoped_repositories: Optional list of repositories in scope (for including zero-vuln repos)
        """
        self.data = None
        self.stats = {}
        self.analytics = {}
        self.trends = {}
        self.scoped_repositories = scoped_repositories or []
    
    def load_vulnerability_data(self, data_source: str) -> bool:
        """
        Load vulnerability data from JSON or CSV file.
        
        Args:
            data_source: Path to JSON or CSV file
            
        Returns:
            True if data loaded successfully
        """
        try:
            if data_source.endswith('.json'):
                with open(data_source, 'r') as f:
                    raw_data = json.load(f)
                self.data = pd.DataFrame(raw_data)
            elif data_source.endswith('.csv'):
                self.data = pd.read_csv(data_source)
            else:
                print(f"❌ Unsupported file format: {data_source}")
                return False
            
            print(f"✅ Loaded {len(self.data)} vulnerability records")
            self._calculate_statistics()
            return True
            
        except Exception as e:
            print(f"❌ Error loading data: {e}")
            return False
    
    def _calculate_statistics(self):
        """Calculate comprehensive vulnerability statistics and analytics."""
        if self.data is None or self.data.empty:
            return
        
        # Basic statistics
        resolved_vulns = self.data[self.data['alert_state'].isin(['fixed', 'dismissed'])]
        fixed_vulns = self.data[self.data['alert_state'] == 'fixed']
        dismissed_vulns = self.data[self.data['alert_state'] == 'dismissed']
        open_vulns = self.data[self.data['alert_state'] == 'open']
        
        # Calculate resolution times for fixed vulnerabilities
        resolution_times = []
        if 'days_to_resolution' in self.data.columns:
            resolution_times = self.data[self.data['days_to_resolution'].notna()]['days_to_resolution'].tolist()
        
        # Calculate vulnerability aging
        vulnerability_ages = []
        if 'vulnerability_age_days' in self.data.columns:
            vulnerability_ages = self.data[self.data['vulnerability_age_days'].notna()]['vulnerability_age_days'].tolist()
        
        # Calculate unique repositories considering scoped repositories
        unique_repos = self.data['repository'].nunique() if not self.data.empty else 0
        if self.scoped_repositories:
            # In scoped mode, count all scoped repositories (including zero-vuln ones)
            unique_repos = len(self.scoped_repositories)
        
        self.stats = {
            # Basic counts
            'total_vulnerabilities': len(self.data),
            'unique_repositories': unique_repos,
            'open_vulnerabilities': len(open_vulns),
            'resolved_vulnerabilities': len(resolved_vulns),
            'fixed_vulnerabilities': len(fixed_vulns),
            'dismissed_vulnerabilities': len(dismissed_vulns),
            
            # Severity breakdown - handle both uppercase and lowercase severity values
            'severity_breakdown': self.data['severity'].value_counts().to_dict(),
            'critical_open': len(open_vulns[open_vulns['severity'].str.upper() == 'CRITICAL']),
            'high_open': len(open_vulns[open_vulns['severity'].str.upper() == 'HIGH']),
            'medium_open': len(open_vulns[open_vulns['severity'].str.upper() == 'MEDIUM']),
            'low_open': len(open_vulns[open_vulns['severity'].str.upper() == 'LOW']),
            
            # Risk metrics
            'average_cvss': round(self.data['cvss_score'].mean(), 2) if 'cvss_score' in self.data.columns else 0,
            'max_cvss': self.data['cvss_score'].max() if 'cvss_score' in self.data.columns else 0,
            'total_risk_score': self._calculate_total_risk_score(),
            
            # Resolution metrics
            'resolution_rate': round((len(resolved_vulns) / len(self.data)) * 100, 1) if len(self.data) > 0 else 0,
            'fix_rate': round((len(fixed_vulns) / len(self.data)) * 100, 1) if len(self.data) > 0 else 0,
            'dismissal_rate': round((len(dismissed_vulns) / len(self.data)) * 100, 1) if len(self.data) > 0 else 0,
            'average_resolution_days': round(np.mean(resolution_times), 1) if resolution_times else 0,
            'median_resolution_days': round(np.median(resolution_times), 1) if resolution_times else 0,
            'average_age_days': round(np.mean(vulnerability_ages), 1) if vulnerability_ages else 0,
            
            # Repository analysis
            'repositories_with_critical': len(self.data[self.data['severity'] == 'CRITICAL']['repository'].unique()),
            'repositories_with_high': len(self.data[self.data['severity'] == 'HIGH']['repository'].unique()),
            'most_vulnerable_repo': self._get_most_vulnerable_repository(),
        }
        
        # Calculate advanced analytics
        self._calculate_advanced_analytics()
        
        # Calculate trends (if historical data available)
        self._calculate_trends()

    def _calculate_total_risk_score(self) -> float:
        """Calculate total organizational risk score."""
        if self.data is None or self.data.empty:
            return 0.0
        
        total_score = 0
        for _, vuln in self.data.iterrows():
            severity_weight = self.SEVERITY_WEIGHTS.get(vuln['severity'], 1)
            # Only count open vulnerabilities for current risk
            if vuln['alert_state'] == 'open':
                total_score += severity_weight
        
        return total_score

    def _get_most_vulnerable_repository(self) -> str:
        """Identify the repository with highest risk score."""
        if self.data is None or self.data.empty:
            return "N/A"
        
        repo_scores = {}
        for repo in self.data['repository'].unique():
            repo_data = self.data[self.data['repository'] == repo]
            repo_score = 0
            for _, vuln in repo_data.iterrows():
                if vuln['alert_state'] == 'open':
                    repo_score += self.SEVERITY_WEIGHTS.get(vuln['severity'], 1)
            repo_scores[repo] = repo_score
        
        if repo_scores:
            return max(repo_scores, key=repo_scores.get)
        return "N/A"

    def _calculate_advanced_analytics(self):
        """Calculate advanced analytics and insights."""
        if self.data is None or self.data.empty:
            return
        
        # Vulnerability distribution analysis
        repo_vuln_counts = self.data['repository'].value_counts()
        
        # Package analysis
        package_analysis = {}
        if 'package_name' in self.data.columns:
            package_vulns = self.data['package_name'].value_counts()
            package_analysis = {
                'most_vulnerable_package': package_vulns.index[0] if len(package_vulns) > 0 else "N/A",
                'packages_with_multiple_vulns': len(package_vulns[package_vulns > 1]),
                'unique_vulnerable_packages': len(package_vulns)
            }
        
        # Ecosystem analysis
        ecosystem_analysis = {}
        if 'package_ecosystem' in self.data.columns:
            ecosystem_vulns = self.data['package_ecosystem'].value_counts()
            ecosystem_analysis = {
                'most_vulnerable_ecosystem': ecosystem_vulns.index[0] if len(ecosystem_vulns) > 0 else "N/A",
                'ecosystem_distribution': ecosystem_vulns.to_dict()
            }
        
        self.analytics = {
            'repository_analysis': {
                'total_repositories': len(repo_vuln_counts),
                'repositories_with_single_vuln': len(repo_vuln_counts[repo_vuln_counts == 1]),
                'repositories_with_multiple_vulns': len(repo_vuln_counts[repo_vuln_counts > 1]),
                'average_vulns_per_repo': round(repo_vuln_counts.mean(), 1),
                'max_vulns_in_single_repo': repo_vuln_counts.max(),
                'repo_risk_distribution': self._calculate_repo_risk_distribution()
            },
            'package_analysis': package_analysis,
            'ecosystem_analysis': ecosystem_analysis,
            'severity_analysis': self._calculate_severity_analysis(),
            'aging_analysis': self._calculate_aging_analysis()
        }

    def _calculate_repo_risk_distribution(self) -> Dict:
        """Calculate risk distribution across repositories."""
        risk_levels = {'low': 0, 'medium': 0, 'high': 0, 'critical': 0}
        
        for repo in self.data['repository'].unique():
            repo_data = self.data[
                (self.data['repository'] == repo) & 
                (self.data['alert_state'] == 'open')
            ]
            repo_score = sum(self.SEVERITY_WEIGHTS.get(vuln['severity'], 1) for _, vuln in repo_data.iterrows())
            
            if repo_score >= 50:
                risk_levels['critical'] += 1
            elif repo_score >= 20:
                risk_levels['high'] += 1
            elif repo_score >= 5:
                risk_levels['medium'] += 1
            else:
                risk_levels['low'] += 1
        
        return risk_levels

    def _calculate_severity_analysis(self) -> Dict:
        """Analyze severity patterns and trends."""
        open_vulns = self.data[self.data['alert_state'] == 'open']
        
        return {
            'severity_priority_score': self._calculate_priority_score(),
            'open_by_severity': open_vulns['severity'].value_counts().to_dict(),
            'resolution_by_severity': self._calculate_resolution_by_severity(),
            'average_cvss_by_severity': self._calculate_avg_cvss_by_severity()
        }

    def _calculate_priority_score(self) -> float:
        """Calculate overall priority score for open vulnerabilities."""
        open_vulns = self.data[self.data['alert_state'] == 'open']
        if open_vulns.empty:
            return 0.0
        
        total_score = sum(self.SEVERITY_WEIGHTS.get(vuln['severity'], 1) for _, vuln in open_vulns.iterrows())
        return round(total_score / len(open_vulns), 2)

    def _calculate_resolution_by_severity(self) -> Dict:
        """Calculate resolution rates by severity level."""
        resolution_rates = {}
        for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
            severity_vulns = self.data[self.data['severity'] == severity]
            if len(severity_vulns) > 0:
                resolved = len(severity_vulns[severity_vulns['alert_state'].isin(['fixed', 'dismissed'])])
                resolution_rates[severity] = round((resolved / len(severity_vulns)) * 100, 1)
            else:
                resolution_rates[severity] = 0.0
        return resolution_rates

    def _calculate_avg_cvss_by_severity(self) -> Dict:
        """Calculate average CVSS scores by severity level."""
        if 'cvss_score' not in self.data.columns:
            return {}
        
        return self.data.groupby('severity')['cvss_score'].mean().round(2).to_dict()

    def _calculate_aging_analysis(self) -> Dict:
        """Analyze vulnerability aging patterns."""
        if 'vulnerability_age_days' not in self.data.columns:
            return {}
        
        open_vulns = self.data[self.data['alert_state'] == 'open']
        ages = open_vulns['vulnerability_age_days'].dropna()
        
        if ages.empty:
            return {}
        
        return {
            'average_age': round(ages.mean(), 1),
            'median_age': round(ages.median(), 1),
            'oldest_vulnerability': ages.max(),
            'newest_vulnerability': ages.min(),
            'aged_30_days': len(ages[ages >= 30]),
            'aged_60_days': len(ages[ages >= 60]),
            'aged_90_days': len(ages[ages >= 90]),
            'stale_vulnerabilities': len(ages[ages >= 180])  # 6 months+
        }

    def _calculate_trends(self):
        """Calculate trend analysis (placeholder for historical data)."""
        # This would be enhanced with historical data comparison
        self.trends = {
            'trend_analysis_available': False,
            'note': 'Historical data required for trend analysis',
            'recommendations': [
                'Implement regular scanning to build historical trends',
                'Track resolution velocity over time',
                'Monitor new vulnerability introduction rates'
            ]
        }
    
    def generate_repository_executive_summary(self) -> pd.DataFrame:
        """
        Generate original repository-focused executive summary report.
        Includes all scoped repositories, even those with zero vulnerabilities.
        
        Returns:
            DataFrame with repository-based executive summary data
        """
        if self.data is None or self.data.empty:
            # If we have scoped repositories but no vulnerability data, show all as clean
            if self.scoped_repositories:
                print("📊 Generating repository-focused executive summary for scoped repositories...")
                print("⚠️  No vulnerability data found, showing all scoped repositories as clean")
                repo_summary = []
                for repo in self.scoped_repositories:
                    repo_summary.append({
                        'Repository Name': repo,
                        'Risk Score': 0,
                        'Total Open': 0,
                        'Critical': 0,
                        'High': 0,
                        'Medium': 0,
                        'Low': 0,
                        'Total All Issues': 0,
                        'Resolved Issues': 0,
                        'Avg Resolution Days': 'N/A',
                        'Oldest Open (Days)': 'N/A'
                    })
                summary_df = pd.DataFrame(repo_summary)
                summary_df.insert(0, 'Priority Rank', range(1, len(summary_df) + 1))
                print(f"✅ Repository executive summary generated for {len(summary_df)} repositories (all clean)")
                return summary_df
            return pd.DataFrame()
        
        print("📊 Generating repository-focused executive summary...")
        
        # Debug: Check what alert states exist
        print(f"  Debug - Alert states in data: {self.data['alert_state'].value_counts().to_dict()}")
        
        # Filter for open vulnerabilities only
        open_vulns = self.data[self.data['alert_state'] == 'open'].copy()
        
        print(f"  Debug - Found {len(open_vulns)} open vulnerabilities")
        if not open_vulns.empty:
            print(f"  Debug - Severity distribution: {open_vulns['severity'].value_counts().to_dict()}")
        
        # Determine which repositories to include
        if self.scoped_repositories:
            # Include all scoped repositories
            repositories_to_include = self.scoped_repositories
            print(f"  Debug - Including all {len(repositories_to_include)} scoped repositories")
        else:
            # Only include repositories with open vulnerabilities
            if open_vulns.empty:
                print("⚠️  No open vulnerabilities found")
                return pd.DataFrame()
            repositories_to_include = open_vulns['repository'].unique().tolist()
            print(f"  Debug - Including {len(repositories_to_include)} repositories with vulnerabilities")
        
        # Group by repository and calculate metrics
        repo_summary = []
        
        for repo in repositories_to_include:
            repo_vulns = open_vulns[open_vulns['repository'] == repo] if not open_vulns.empty else pd.DataFrame()
            
            if not repo_vulns.empty:
                # Normalize severity values to uppercase
                repo_vulns['severity'] = repo_vulns['severity'].str.upper()
                
                # Count by severity
                severity_counts = repo_vulns['severity'].value_counts()
                
                # Debug: Print severity counts for first few repos
                if len(repo_summary) < 3:
                    print(f"  Debug - {repo}: severities = {severity_counts.to_dict()}")
                
                # Calculate risk score
                risk_score = sum(
                    severity_counts.get(severity, 0) * weight
                    for severity, weight in self.SEVERITY_WEIGHTS.items()
                )
                
                # All vulnerabilities for this repo (including closed)
                all_repo_vulns = self.data[self.data['repository'] == repo]
                resolved_vulns = all_repo_vulns[all_repo_vulns['alert_state'].isin(['fixed', 'dismissed'])]
                avg_resolution_days = resolved_vulns['days_to_resolution'].mean() if 'days_to_resolution' in resolved_vulns.columns and len(resolved_vulns) > 0 else None
                
                repo_summary.append({
                    'Repository Name': repo,
                    'Risk Score': risk_score,
                    'Total Open': len(repo_vulns),
                    'Critical': severity_counts.get('CRITICAL', 0),
                    'High': severity_counts.get('HIGH', 0),
                    'Medium': severity_counts.get('MEDIUM', 0),
                    'Low': severity_counts.get('LOW', 0),
                    'Total All Issues': len(all_repo_vulns),
                    'Resolved Issues': len(resolved_vulns),
                    'Avg Resolution Days': round(avg_resolution_days, 1) if avg_resolution_days else 'N/A',
                    'Oldest Open (Days)': repo_vulns['vulnerability_age_days'].max() if len(repo_vulns) > 0 and 'vulnerability_age_days' in repo_vulns.columns else 'N/A'
                })
            else:
                # Repository with zero vulnerabilities
                all_repo_vulns = self.data[self.data['repository'] == repo] if not self.data.empty else pd.DataFrame()
                resolved_vulns = all_repo_vulns[all_repo_vulns['alert_state'].isin(['fixed', 'dismissed'])] if not all_repo_vulns.empty else pd.DataFrame()
                avg_resolution_days = resolved_vulns['days_to_resolution'].mean() if 'days_to_resolution' in resolved_vulns.columns and len(resolved_vulns) > 0 else None
                
                repo_summary.append({
                    'Repository Name': repo,
                    'Risk Score': 0,
                    'Total Open': 0,
                    'Critical': 0,
                    'High': 0,
                    'Medium': 0,
                    'Low': 0,
                    'Total All Issues': len(all_repo_vulns),
                    'Resolved Issues': len(resolved_vulns),
                    'Avg Resolution Days': round(avg_resolution_days, 1) if avg_resolution_days else 'N/A',
                    'Oldest Open (Days)': 'N/A'
                })
        
        # Create DataFrame and sort by risk score
        summary_df = pd.DataFrame(repo_summary)
        summary_df = summary_df.sort_values('Risk Score', ascending=False).reset_index(drop=True)
        
        # Add priority ranking
        summary_df.insert(0, 'Priority Rank', range(1, len(summary_df) + 1))
        
        clean_repos = len(summary_df[summary_df['Total Open'] == 0])
        vuln_repos = len(summary_df[summary_df['Total Open'] > 0])
        
        print(f"✅ Repository executive summary generated for {len(summary_df)} repositories")
        if self.scoped_repositories:
            print(f"   📊 Repositories with vulnerabilities: {vuln_repos}")
            print(f"   ✅ Clean repositories (zero vulnerabilities): {clean_repos}")
        
        return summary_df

    def generate_executive_summary(self) -> pd.DataFrame:
        """
        Generate enhanced executive summary report with comprehensive analytics.
        
        Returns:
            DataFrame with executive summary data including KPIs and trends
        """
        if self.data is None or self.data.empty:
            return pd.DataFrame()
        
        print("📊 Generating enhanced executive summary report...")
        
        # Create comprehensive executive summary
        summary_data = []
        
        # Overall organizational metrics
        summary_data.extend([
            {'Metric': 'ORGANIZATIONAL OVERVIEW', 'Value': '', 'Details': '', 'Status': 'HEADER'},
            {'Metric': 'Total Vulnerabilities', 'Value': self.stats['total_vulnerabilities'], 
             'Details': f"Across {self.stats['unique_repositories']} repositories", 'Status': 'INFO'},
            {'Metric': 'Open Vulnerabilities', 'Value': self.stats['open_vulnerabilities'], 
             'Details': f"{(self.stats['open_vulnerabilities']/self.stats['total_vulnerabilities']*100):.1f}% of total", 
             'Status': 'OPEN' if self.stats['open_vulnerabilities'] > 0 else 'GOOD'},
            {'Metric': 'Resolution Rate', 'Value': f"{self.stats['resolution_rate']}%", 
             'Details': f"Fixed: {self.stats['fix_rate']}%, Dismissed: {self.stats['dismissal_rate']}%", 
             'Status': 'GOOD' if self.stats['resolution_rate'] >= 80 else 'WARNING'},
            {'Metric': 'Total Risk Score', 'Value': int(self.stats['total_risk_score']), 
             'Details': 'Weighted severity score for open vulnerabilities', 
             'Status': 'CRITICAL' if self.stats['total_risk_score'] > 100 else 'WARNING' if self.stats['total_risk_score'] > 50 else 'GOOD'},
        ])
        
        # Severity breakdown
        summary_data.extend([
            {'Metric': '', 'Value': '', 'Details': '', 'Status': ''},
            {'Metric': 'SEVERITY BREAKDOWN', 'Value': '', 'Details': '', 'Status': 'HEADER'},
            {'Metric': 'Critical Vulnerabilities', 'Value': self.stats['critical_open'], 
             'Details': f"Requires immediate attention", 
             'Status': 'CRITICAL' if self.stats['critical_open'] > 0 else 'GOOD'},
            {'Metric': 'High Vulnerabilities', 'Value': self.stats['high_open'], 
             'Details': f"Priority remediation", 
             'Status': 'HIGH' if self.stats['high_open'] > 0 else 'GOOD'},
            {'Metric': 'Medium Vulnerabilities', 'Value': self.stats['medium_open'], 
             'Details': f"Standard remediation timeline", 
             'Status': 'MEDIUM' if self.stats['medium_open'] > 0 else 'GOOD'},
            {'Metric': 'Low Vulnerabilities', 'Value': self.stats['low_open'], 
             'Details': f"Monitor and schedule", 
             'Status': 'LOW' if self.stats['low_open'] > 0 else 'GOOD'},
        ])
        
        # Risk and quality metrics
        summary_data.extend([
            {'Metric': '', 'Value': '', 'Details': '', 'Status': ''},
            {'Metric': 'RISK & QUALITY METRICS', 'Value': '', 'Details': '', 'Status': 'HEADER'},
            {'Metric': 'Average CVSS Score', 'Value': self.stats['average_cvss'], 
             'Details': f"Maximum: {self.stats['max_cvss']}", 
             'Status': 'CRITICAL' if self.stats['average_cvss'] >= 7.0 else 'WARNING' if self.stats['average_cvss'] >= 4.0 else 'GOOD'},
            {'Metric': 'Most Vulnerable Repository', 'Value': self.stats['most_vulnerable_repo'], 
             'Details': 'Highest risk score', 'Status': 'WARNING'},
            {'Metric': 'Repositories with Critical', 'Value': self.stats['repositories_with_critical'], 
             'Details': f"Out of {self.stats['unique_repositories']} total repositories", 
             'Status': 'CRITICAL' if self.stats['repositories_with_critical'] > 0 else 'GOOD'},
        ])
        
        # Performance metrics
        if self.stats['average_resolution_days'] > 0:
            resolution_status = 'GOOD' if self.stats['average_resolution_days'] <= 30 else 'WARNING' if self.stats['average_resolution_days'] <= 60 else 'CRITICAL'
            summary_data.extend([
                {'Metric': '', 'Value': '', 'Details': '', 'Status': ''},
                {'Metric': 'PERFORMANCE METRICS', 'Value': '', 'Details': '', 'Status': 'HEADER'},
                {'Metric': 'Average Resolution Time', 'Value': f"{self.stats['average_resolution_days']} days", 
                 'Details': f"Median: {self.stats['median_resolution_days']} days", 'Status': resolution_status},
                {'Metric': 'Average Vulnerability Age', 'Value': f"{self.stats['average_age_days']} days", 
                 'Details': 'For currently open vulnerabilities', 
                 'Status': 'GOOD' if self.stats['average_age_days'] <= 30 else 'WARNING' if self.stats['average_age_days'] <= 90 else 'CRITICAL'},
            ])
        
        # Advanced analytics insights
        if self.analytics:
            repo_analysis = self.analytics.get('repository_analysis', {})
            aging_analysis = self.analytics.get('aging_analysis', {})
            
            summary_data.extend([
                {'Metric': '', 'Value': '', 'Details': '', 'Status': ''},
                {'Metric': 'ADVANCED INSIGHTS', 'Value': '', 'Details': '', 'Status': 'HEADER'},
                {'Metric': 'Avg Vulnerabilities per Repo', 'Value': repo_analysis.get('average_vulns_per_repo', 0), 
                 'Details': f"Max in single repo: {repo_analysis.get('max_vulns_in_single_repo', 0)}", 'Status': 'INFO'},
            ])
            
            if aging_analysis:
                stale_vulns = aging_analysis.get('stale_vulnerabilities', 0)
                summary_data.append({
                    'Metric': 'Stale Vulnerabilities (180+ days)', 'Value': stale_vulns, 
                    'Details': 'May require special attention', 
                    'Status': 'CRITICAL' if stale_vulns > 0 else 'GOOD'
                })
        
        # Recommendations
        recommendations = self._generate_recommendations()
        if recommendations:
            summary_data.extend([
                {'Metric': '', 'Value': '', 'Details': '', 'Status': ''},
                {'Metric': 'KEY RECOMMENDATIONS', 'Value': '', 'Details': '', 'Status': 'HEADER'},
            ])
            for i, rec in enumerate(recommendations[:5], 1):  # Top 5 recommendations
                summary_data.append({
                    'Metric': f'Priority {i}', 'Value': rec['action'], 
                    'Details': rec['reason'], 'Status': rec['priority']
                })
        
        return pd.DataFrame(summary_data)

    def _generate_recommendations(self) -> List[Dict]:
        """Generate intelligent recommendations based on vulnerability analysis."""
        recommendations = []
        
        # Critical vulnerabilities
        if self.stats['critical_open'] > 0:
            recommendations.append({
                'action': f"Address {self.stats['critical_open']} Critical Vulnerabilities",
                'reason': "Critical severity requires immediate attention",
                'priority': 'CRITICAL'
            })
        
        # High vulnerabilities
        if self.stats['high_open'] > 5:
            recommendations.append({
                'action': f"Prioritize {self.stats['high_open']} High Severity Issues",
                'reason': "Large number of high-severity vulnerabilities",
                'priority': 'HIGH'
            })
        
        # Resolution rate
        if self.stats['resolution_rate'] < 70:
            recommendations.append({
                'action': "Improve Vulnerability Resolution Process",
                'reason': f"Current resolution rate is {self.stats['resolution_rate']}%",
                'priority': 'MEDIUM'
            })
        
        # Aging vulnerabilities
        if self.analytics and self.analytics.get('aging_analysis', {}).get('stale_vulnerabilities', 0) > 0:
            stale_count = self.analytics['aging_analysis']['stale_vulnerabilities']
            recommendations.append({
                'action': f"Review {stale_count} Stale Vulnerabilities",
                'reason': "Vulnerabilities older than 180 days need review",
                'priority': 'MEDIUM'
            })
        
        # Repository concentration
        if self.analytics and self.analytics.get('repository_analysis', {}).get('max_vulns_in_single_repo', 0) > 20:
            max_vulns = self.analytics['repository_analysis']['max_vulns_in_single_repo']
            recommendations.append({
                'action': f"Focus on High-Risk Repository",
                'reason': f"One repository has {max_vulns} vulnerabilities",
                'priority': 'HIGH'
            })
        
        # Package ecosystem focus
        if self.analytics and self.analytics.get('package_analysis', {}).get('packages_with_multiple_vulns', 0) > 0:
            multi_vuln_packages = self.analytics['package_analysis']['packages_with_multiple_vulns']
            recommendations.append({
                'action': "Review Packages with Multiple Vulnerabilities",
                'reason': f"{multi_vuln_packages} packages have multiple vulnerabilities",
                'priority': 'MEDIUM'
            })
        
        return recommendations

    def generate_detailed_report(self) -> pd.DataFrame:
        """
        Generate detailed technical vulnerability report.
        
        Returns:
            DataFrame with detailed vulnerability data
        """
        if self.data is None or self.data.empty:
            return pd.DataFrame()
        
        print("🔍 Generating detailed technical report...")
        
        # Select and rename columns for the detailed report
        detailed_columns = {
            'repository': 'Repository Name',
            'alert_state': 'Status',
            'current_status': 'Detailed Status',
            'severity': 'Severity',
            'cvss_score': 'CVSS Score',
            'summary': 'Vulnerability Title',
            'package_name': 'Component',
            'current_version': 'Existing Version',
            'vulnerable_version_range': 'Vulnerable Versions',
            'first_patched_version': 'Patched Version',
            'manifest_path': 'Detected In',
            'alert_created_date': 'Detected On',
            'alert_fixed_date': 'Fixed On',
            'alert_dismissed_date': 'Dismissed On',
            'resolution_method': 'Resolution Method',
            'days_to_resolution': 'Days to Resolution',
            'vulnerability_age_days': 'Age (Days)',
            'dismisser_login': 'Dismissed By',
            'alert_dismissed_reason': 'Dismissal Reason',
            'cvss_vector': 'CVSS Vector',
            'cve_id': 'CVE ID',
            'ghsa_id': 'GHSA ID',
            'alert_url': 'Alert URL',
            'description': 'Description'
        }
        
        # Create detailed report
        detailed_df = self.data.copy()
        
        # Ensure all required columns exist
        for col in detailed_columns.keys():
            if col not in detailed_df.columns:
                detailed_df[col] = 'N/A'
        
        # Select and rename columns
        detailed_df = detailed_df[list(detailed_columns.keys())].rename(columns=detailed_columns)
        
        # Clean and format data
        detailed_df['Status'] = detailed_df['Status'].str.upper()
        detailed_df['Severity'] = detailed_df['Severity'].str.upper()
        detailed_df['CVSS Score'] = pd.to_numeric(detailed_df['CVSS Score'], errors='coerce').fillna(0.0)
        
        # Format dates - these are already formatted by the scanner
        date_columns = ['Detected On', 'Fixed On', 'Dismissed On']
        for col in date_columns:
            if col in detailed_df.columns:
                # Dates are already formatted, just ensure consistent display
                detailed_df[col] = detailed_df[col].fillna('N/A')
        
        # Sort by severity and CVSS score
        severity_order = ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']
        detailed_df['severity_rank'] = detailed_df['Severity'].map({s: i for i, s in enumerate(severity_order)})
        detailed_df = detailed_df.sort_values(['severity_rank', 'CVSS Score'], ascending=[True, False])
        detailed_df = detailed_df.drop('severity_rank', axis=1).reset_index(drop=True)
        
        print(f"✅ Detailed report generated with {len(detailed_df)} vulnerabilities")
        return detailed_df
    
    def _escape_excel_formulas(self, value) -> str:
        """
        Escape potential Excel formulas to prevent security warnings.
        
        Args:
            value: Value to escape
            
        Returns:
            Escaped value safe for Excel
        """
        if pd.isna(value):
            return value
        
        str_value = str(value)
        if str_value.startswith(('=', '+', '-', '@')):
            return "'" + str_value
        return str_value
    
    def _escape_dataframe_formulas(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply Excel formula escaping to all string columns in a DataFrame.
        
        Args:
            df: DataFrame to escape
            
        Returns:
            DataFrame with escaped formulas
        """
        if df.empty:
            return df
        
        df_copy = df.copy()
        for col in df_copy.columns:
            if df_copy[col].dtype == 'object':  # String columns
                df_copy[col] = df_copy[col].apply(self._escape_excel_formulas)
        return df_copy
    
    def _apply_excel_formatting(self, worksheet, dataframe: pd.DataFrame, report_type: str):
        """
        Apply professional Excel formatting to worksheet with enhanced styling.
        
        Args:
            worksheet: openpyxl worksheet object
            dataframe: DataFrame being formatted
            report_type: 'executive', 'enhanced', 'detailed', or 'dashboard'
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
        except ImportError as e:
            print(f"⚠️  Excel formatting import error: {e}")
            return
        
        try:
            # Define color schemes for different report types
            color_schemes = {
                'executive': {'header_color': '366092', 'accent_color': '4472C4'},
                'enhanced': {'header_color': '2E8B57', 'accent_color': '32CD32'},  # Sea Green
                'detailed': {'header_color': '4472C4', 'accent_color': '366092'},
                'dashboard': {'header_color': 'FF6B35', 'accent_color': 'F7931E'}  # Orange
            }
            
            scheme = color_schemes.get(report_type, color_schemes['executive'])
            
            # Header formatting
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_fill = PatternFill(
                start_color=scheme['header_color'],
                end_color=scheme['header_color'],
                fill_type='solid'
            )
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply header formatting
            for col_num in range(1, len(dataframe.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Enhanced formatting for different report types
            if report_type == 'enhanced':
                self._apply_enhanced_kpi_formatting(worksheet, dataframe)
            elif report_type == 'executive':
                self._apply_executive_formatting(worksheet, dataframe)
            elif report_type == 'dashboard':
                self._apply_dashboard_formatting(worksheet, dataframe)
            elif report_type == 'detailed':
                self._apply_detailed_formatting(worksheet, dataframe)
            
            # Auto-adjust column widths
            for col_idx, column in enumerate(worksheet.columns):
                column_letter = get_column_letter(col_idx + 1)
                column_name = dataframe.columns[col_idx] if col_idx < len(dataframe.columns) else ""
                
                # Calculate column width
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set width based on column type and content
                if 'description' in column_name.lower():
                    width = 120
                elif 'details' in column_name.lower():
                    width = 80
                elif 'status' in column_name.lower():
                    width = 15
                elif 'value' in column_name.lower():
                    width = 20
                else:
                    width = min(max(max_length + 2, 12), 50)
                
                worksheet.column_dimensions[column_letter].width = width
                
        except ImportError:
            print("⚠️  openpyxl not available for advanced formatting")
        except Exception as e:
            print(f"⚠️  Formatting error: {e}")

    def _apply_enhanced_kpi_formatting(self, worksheet, dataframe: pd.DataFrame):
        """Apply special formatting for enhanced KPI summary."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Format header rows and status indicators
            for row_num in range(2, len(dataframe) + 2):
                metric_cell = worksheet.cell(row=row_num, column=1)  # Metric column
                status_cell = worksheet.cell(row=row_num, column=4) if len(dataframe.columns) >= 4 else None  # Status column
                
                # Header row formatting
                if 'HEADER' in str(dataframe.iloc[row_num-2, -1]) if len(dataframe) >= row_num-1 else False:
                    for col_num in range(1, len(dataframe.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.font = Font(bold=True, color='2E8B57', size=11)
                        cell.fill = PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')
                
                # Status-based color coding
                if status_cell and status_cell.value:
                    status_colors = {
                        'CRITICAL': {'bg': 'FFE6E6', 'font': 'CC0000'},
                        'WARNING': {'bg': 'FFF3CD', 'font': 'B45309'},
                        'GOOD': {'bg': 'E6F3E6', 'font': '155724'},
                        'INFO': {'bg': 'E7F3FF', 'font': '0C5460'}
                    }
                    
                    status_value = str(status_cell.value).upper()
                    if status_value in status_colors:
                        colors = status_colors[status_value]
                        for col_num in range(1, len(dataframe.columns) + 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == '00000000':  # Don't override header formatting
                                cell.fill = PatternFill(start_color=colors['bg'], end_color=colors['bg'], fill_type='solid')
                                if not cell.font.bold:  # Don't override header font
                                    cell.font = Font(color=colors['font'])
                
        except Exception as e:
            print(f"⚠️  Enhanced KPI formatting error: {e}")

    def _apply_executive_formatting(self, worksheet, dataframe: pd.DataFrame):
        """Apply formatting for repository executive summary."""
        try:
            from openpyxl.styles import Font, PatternFill
            from openpyxl.formatting.rule import ColorScaleRule
            
            # Apply risk score color scale
            if 'Risk Score' in dataframe.columns:
                risk_col_idx = list(dataframe.columns).index('Risk Score') + 1
                risk_col_letter = get_column_letter(risk_col_idx)
                
                # Color scale from green (low risk) to red (high risk)
                color_scale = ColorScaleRule(
                    start_type='min', start_color='90EE90',
                    end_type='max', end_color='FF6B6B'
                )
                worksheet.conditional_formatting.add(f'{risk_col_letter}2:{risk_col_letter}{len(dataframe)+1}', color_scale)
            
            # Highlight top 5 highest risk repositories
            for row_num in range(2, min(7, len(dataframe) + 2)):  # Top 5 rows
                for col_num in range(1, len(dataframe.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if row_num <= 3:  # Top 3 - red highlight
                        cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                    elif row_num <= 6:  # Top 4-5 - yellow highlight
                        cell.fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
                        
        except Exception as e:
            print(f"⚠️  Executive formatting error: {e}")

    def _apply_dashboard_formatting(self, worksheet, dataframe: pd.DataFrame):
        """Apply formatting for dashboard summary."""
        try:
            from openpyxl.styles import Font, PatternFill
            
            # Apply status-based formatting
            if 'Status' in dataframe.columns:
                status_col_idx = list(dataframe.columns).index('Status') + 1
                
                for row_num in range(2, len(dataframe) + 2):
                    status_cell = worksheet.cell(row=row_num, column=status_col_idx)
                    if status_cell.value:
                        status_value = str(status_cell.value)
                        if '🔴' in status_value:
                            for col_num in range(1, len(dataframe.columns) + 1):
                                worksheet.cell(row=row_num, column=col_num).fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                        elif '🟡' in status_value:
                            for col_num in range(1, len(dataframe.columns) + 1):
                                worksheet.cell(row=row_num, column=col_num).fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
                        elif '🟢' in status_value:
                            for col_num in range(1, len(dataframe.columns) + 1):
                                worksheet.cell(row=row_num, column=col_num).fill = PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')
                            
        except Exception as e:
            print(f"⚠️  Dashboard formatting error: {e}")

    def _apply_detailed_formatting(self, worksheet, dataframe: pd.DataFrame):
        """Apply formatting for detailed vulnerabilities."""
        try:
            from openpyxl.styles import Font, PatternFill
            
            # Apply severity color coding
            if 'Severity' in dataframe.columns:
                severity_col_idx = list(dataframe.columns).index('Severity') + 1
                
                for row_num in range(2, len(dataframe) + 2):
                    severity_cell = worksheet.cell(row=row_num, column=severity_col_idx)
                    if severity_cell.value:
                        severity = str(severity_cell.value).upper()
                        if severity in self.SEVERITY_COLORS:
                            colors = self.SEVERITY_COLORS[severity]
                            severity_cell.fill = PatternFill(start_color=colors['bg'], end_color=colors['bg'], fill_type='solid')
                            severity_cell.font = Font(color=colors['font'], bold=True)
            
            # Apply status color coding
            if 'Status' in dataframe.columns:
                status_col_idx = list(dataframe.columns).index('Status') + 1
                
                for row_num in range(2, len(dataframe) + 2):
                    status_cell = worksheet.cell(row=row_num, column=status_col_idx)
                    if status_cell.value:
                        status = str(status_cell.value).upper()
                        if status in self.STATUS_COLORS:
                            colors = self.STATUS_COLORS[status.lower()]
                            status_cell.fill = PatternFill(start_color=colors['bg'], end_color=colors['bg'], fill_type='solid')
                            status_cell.font = Font(color=colors['font'])
                            
        except Exception as e:
            print(f"⚠️  Detailed formatting error: {e}")

    def save_reports(self, repo_executive_df: pd.DataFrame, enhanced_executive_df: pd.DataFrame, 
                    detailed_df: pd.DataFrame, output_dir: Optional[str] = None) -> str:
        """
        Save comprehensive reports with multiple sheets and enhanced formatting.
        
        Args:
            repo_executive_df: Repository-focused executive summary DataFrame
            enhanced_executive_df: Enhanced KPI-focused executive summary DataFrame
            detailed_df: Detailed vulnerabilities DataFrame
            output_dir: Optional custom output directory
            
        Returns:
            Path to the generated reports folder
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create organized folder structure
        if output_dir:
            reports_dir = Path(output_dir)
        else:
            reports_dir = Path("reports")
        
        timestamped_dir = reports_dir / f"security_reports_{timestamp}"
        timestamped_dir.mkdir(parents=True, exist_ok=True)
        
        print(f"📁 Creating comprehensive reports in: {timestamped_dir}")
        
        # Save CSV files (use repository executive summary for main CSV)
        if not repo_executive_df.empty:
            repo_executive_df.to_csv(timestamped_dir / "executive_summary.csv", index=False)
            print("✅ Executive CSV: executive_summary.csv")
        
        if not enhanced_executive_df.empty:
            enhanced_executive_df.to_csv(timestamped_dir / "executive_kpi_summary.csv", index=False)
            print("✅ Enhanced KPI CSV: executive_kpi_summary.csv")
        
        if not detailed_df.empty:
            detailed_df.to_csv(timestamped_dir / "detailed_vulnerabilities.csv", index=False)
            print("✅ Detailed CSV: detailed_vulnerabilities.csv")
        
        # Save comprehensive Excel file with multiple sheets
        self._save_comprehensive_excel_reports(repo_executive_df, enhanced_executive_df, detailed_df, timestamped_dir)
        
        # Create README
        self._create_readme(repo_executive_df, detailed_df, timestamped_dir)
        
        print(f"✅ Comprehensive report suite generated successfully in: {timestamped_dir}")
        return str(timestamped_dir)
    
    def _save_comprehensive_excel_reports(self, repo_executive_df: pd.DataFrame, 
                                         enhanced_executive_df: pd.DataFrame, detailed_df: pd.DataFrame, output_dir: Path):
        """Save comprehensive Excel reports with multiple sheets and enhanced formatting."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
            
            # Apply formula escaping to all DataFrames before Excel export
            safe_repo_exec_df = self._escape_dataframe_formulas(repo_executive_df)
            safe_enhanced_df = self._escape_dataframe_formulas(enhanced_executive_df)
            safe_detailed_df = self._escape_dataframe_formulas(detailed_df)
            
            # Create comprehensive executive summary workbook
            exec_filename = output_dir / "executive_summary.xlsx"
            with pd.ExcelWriter(exec_filename, engine='openpyxl') as writer:
                
                # Sheet 1: Repository Executive Summary (Main)
                if not safe_repo_exec_df.empty:
                    safe_repo_exec_df.to_excel(writer, sheet_name='Executive Summary', index=False)
                    ws_exec = writer.sheets['Executive Summary']
                    self._apply_excel_formatting(ws_exec, safe_repo_exec_df, 'executive')
                    print("✅ Executive Summary sheet with repository focus")
                
                # Sheet 2: Enhanced KPI Analytics
                if not safe_enhanced_df.empty:
                    safe_enhanced_df.to_excel(writer, sheet_name='Enhanced Analytics', index=False)
                    ws_enhanced = writer.sheets['Enhanced Analytics']
                    self._apply_excel_formatting(ws_enhanced, safe_enhanced_df, 'enhanced')
                    print("✅ Enhanced Analytics sheet with KPIs and metrics")
                
                # Sheet 3: Quick Dashboard (Summary stats)
                dashboard_data = self._create_dashboard_data()
                if dashboard_data:
                    dashboard_df = pd.DataFrame(dashboard_data)
                    safe_dashboard_df = self._escape_dataframe_formulas(dashboard_df)
                    safe_dashboard_df.to_excel(writer, sheet_name='Dashboard', index=False)
                    ws_dashboard = writer.sheets['Dashboard']
                    self._apply_excel_formatting(ws_dashboard, safe_dashboard_df, 'dashboard')
                    print("✅ Dashboard sheet with key metrics")
            
            # Create detailed vulnerabilities workbook
            detailed_filename = output_dir / "detailed_vulnerabilities.xlsx"
            if not safe_detailed_df.empty:
                with pd.ExcelWriter(detailed_filename, engine='openpyxl') as writer:
                    safe_detailed_df.to_excel(writer, sheet_name='All Vulnerabilities', index=False)
                    ws_detailed = writer.sheets['All Vulnerabilities']
                    self._apply_excel_formatting(ws_detailed, safe_detailed_df, 'detailed')
                    
                    # Add filtered sheets for different severities
                    for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
                        severity_vulns = safe_detailed_df[safe_detailed_df['Severity'] == severity] if 'Severity' in safe_detailed_df.columns else pd.DataFrame()
                        if not severity_vulns.empty:
                            severity_vulns.to_excel(writer, sheet_name=f'{severity.title()} Severity', index=False)
                            ws_severity = writer.sheets[f'{severity.title()} Severity']
                            self._apply_excel_formatting(ws_severity, severity_vulns, 'detailed')
                
                print("✅ Detailed vulnerabilities Excel with severity sheets")
            
        except ImportError:
            print("⚠️  openpyxl not available, skipping Excel formatting")
        except Exception as e:
            print(f"⚠️  Excel formatting error: {e}")

    def _create_dashboard_data(self) -> List[Dict]:
        """Create dashboard summary data for quick overview."""
        if not self.stats:
            return []
        
        dashboard_data = [
            {'Metric': 'Total Vulnerabilities', 'Value': self.stats.get('total_vulnerabilities', 0), 'Target': 'Minimize', 'Status': '🔴' if self.stats.get('total_vulnerabilities', 0) > 100 else '🟡' if self.stats.get('total_vulnerabilities', 0) > 50 else '🟢'},
            {'Metric': 'Open Vulnerabilities', 'Value': self.stats.get('open_vulnerabilities', 0), 'Target': 'Minimize', 'Status': '🔴' if self.stats.get('open_vulnerabilities', 0) > 50 else '🟡' if self.stats.get('open_vulnerabilities', 0) > 20 else '🟢'},
            {'Metric': 'Resolution Rate (%)', 'Value': f"{self.stats.get('resolution_rate', 0):.1f}%", 'Target': '> 80%', 'Status': '🟢' if self.stats.get('resolution_rate', 0) >= 80 else '🟡' if self.stats.get('resolution_rate', 0) >= 60 else '🔴'},
            {'Metric': 'Avg Resolution Days', 'Value': self.stats.get('average_resolution_days', 0), 'Target': '< 30 days', 'Status': '🟢' if self.stats.get('average_resolution_days', 999) <= 30 else '🟡' if self.stats.get('average_resolution_days', 999) <= 60 else '🔴'},
            {'Metric': 'Critical Open', 'Value': self.stats.get('critical_open', 0), 'Target': '0', 'Status': '🟢' if self.stats.get('critical_open', 0) == 0 else '🔴'},
            {'Metric': 'High Open', 'Value': self.stats.get('high_open', 0), 'Target': '< 5', 'Status': '🟢' if self.stats.get('high_open', 0) < 5 else '🟡' if self.stats.get('high_open', 0) < 10 else '🔴'},
            {'Metric': 'Repositories Scanned', 'Value': self.stats.get('unique_repositories', 0), 'Target': 'Full Coverage', 'Status': '🟢'},
            {'Metric': 'Average CVSS Score', 'Value': f"{self.stats.get('average_cvss', 0):.1f}", 'Target': '< 4.0', 'Status': '🟢' if self.stats.get('average_cvss', 0) < 4.0 else '🟡' if self.stats.get('average_cvss', 0) < 7.0 else '🔴'},
        ]
        
        return dashboard_data

    def _create_readme(self, executive_df: pd.DataFrame, detailed_df: pd.DataFrame, output_dir: Path):
        """Create comprehensive README file."""
        readme_content = f"""# Security Reports
Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

## 📊 Report Summary

### Security Posture Overview
- **Total Repositories Analyzed**: {self.stats.get('unique_repositories', 0)}
- **Total Vulnerabilities Found**: {self.stats.get('total_vulnerabilities', 0)}
- **Open Vulnerabilities**: {self.stats.get('open_vulnerabilities', 0)}
- **Resolved Vulnerabilities**: {self.stats.get('resolved_vulnerabilities', 0)}
- **Resolution Rate**: {self.stats.get('resolution_rate', 0):.1f}%
- **Average CVSS Score**: {self.stats.get('average_cvss', 0):.1f}

### Files Generated
- executive_summary.xlsx - Multi-sheet executive workbook
- detailed_vulnerabilities.xlsx - Complete vulnerability inventory
- *.csv files - Data exports for analysis

*Generated by Security Vulnerability Scanner v2.1*
"""
        
        with open(output_dir / "README.md", 'w', encoding='utf-8') as f:
            f.write(readme_content)


def main():
    """Example usage of the SecurityReportGenerator."""
    import sys
    
    if len(sys.argv) != 2:
        print("Usage: python security_report_generator.py <vulnerability_data.json>")
        return
    
    data_file = sys.argv[1]
    
    # Initialize generator
    generator = SecurityReportGenerator()
    
    # Load data
    if not generator.load_vulnerability_data(data_file):
        print(f"❌ Failed to load data from {data_file}")
        return
    
    # Generate reports
    repo_executive_df = generator.generate_repository_executive_summary()
    enhanced_executive_df = generator.generate_executive_summary()
    detailed_df = generator.generate_detailed_report()
    
    # Save reports
    output_dir = generator.save_reports(repo_executive_df, enhanced_executive_df, detailed_df)
    print(f"\n🎉 Reports generated successfully in: {output_dir}")


if __name__ == "__main__":
    main()
