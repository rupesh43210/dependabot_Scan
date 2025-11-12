#!/usr/bin/env python3
"""
Code Scanning Report Generator

Professional report generator for GitHub Code Scanning alerts with
advanced analytics and comprehensive visualizations.

Author: GitHub Copilot
Version: 1.0.0
"""

import os
import pandas as pd
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional


class CodeScanningReportGenerator:
    """
    Report generator for Code Scanning alerts.
    
    Features:
    - Executive summary reports with KPIs
    - Detailed technical reports
    - Repository-focused summaries
    - Risk scoring and prioritization
    """
    
    # Risk scoring weights
    SEVERITY_WEIGHTS = {
        'CRITICAL': 50,
        'HIGH': 20,
        'MEDIUM': 5,
        'LOW': 1,
        'ERROR': 20,
        'WARNING': 3,
        'NOTE': 1
    }
    
    def __init__(self, scoped_repositories: Optional[List[str]] = None, active_scope: Optional[str] = None):
        """
        Initialize the report generator.
        
        Args:
            scoped_repositories: Optional list of repositories in scope
            active_scope: Optional name of the active scope
        """
        self.data = None
        self.stats = {}
        self.scoped_repositories = scoped_repositories or []
        self.active_scope = active_scope
        self.repository_metadata = {}
        self.output_dir = self._load_output_dir_from_config()
    
    def _load_output_dir_from_config(self) -> str:
        """Load output directory from config/config.json."""
        try:
            # Get project root (2 levels up from reporters/)
            project_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            config_path = os.path.join(project_root, 'config', 'config.json')
            with open(config_path, 'r') as f:
                config = json.load(f)
            # Use codeql_output_dir if available, otherwise fall back to default
            return config.get('scan', {}).get('codeql_output_dir', './reports/codeql_alerts')
        except Exception:
            return './reports/codeql_alerts'
    
    def load_repository_metadata(self, metadata_file: str) -> bool:
        """Load repository metadata from JSON file."""
        try:
            if os.path.exists(metadata_file):
                with open(metadata_file, 'r') as f:
                    self.repository_metadata = json.load(f)
                print(f"✅ Loaded metadata for {len(self.repository_metadata)} repositories")
                return True
            else:
                print(f"ℹ️  Metadata file not found: {metadata_file}")
                return False
        except Exception as e:
            print(f"⚠️  Error loading repository metadata: {e}")
            return False
    
    def load_alert_data(self, data_source: str) -> bool:
        """
        Load Code Scanning alert data from JSON or CSV file.
        
        Args:
            data_source: Path to JSON or CSV file
            
        Returns:
            True if data loaded successfully
        """
        try:
            if data_source.endswith('.json'):
                with open(data_source, 'r') as f:
                    raw_data = json.load(f)
                self.data = pd.DataFrame(raw_data)
            elif data_source.endswith('.csv'):
                self.data = pd.read_csv(data_source)
            else:
                print(f"❌ Unsupported file format: {data_source}")
                return False
            
            print(f"✅ Loaded {len(self.data)} Code Scanning alert records")
            self._calculate_statistics()
            return True
            
        except Exception as e:
            print(f"❌ Error loading data: {e}")
            return False
    
    def _calculate_statistics(self):
        """Calculate alert statistics."""
        if self.data is None or self.data.empty:
            return
        
        open_alerts = self.data[self.data['alert_state'] == 'open']
        fixed_alerts = self.data[self.data['alert_state'] == 'fixed']
        dismissed_alerts = self.data[self.data['alert_state'] == 'dismissed']
        
        self.stats = {
            'total_alerts': len(self.data),
            'open_alerts': len(open_alerts),
            'fixed_alerts': len(fixed_alerts),
            'dismissed_alerts': len(dismissed_alerts),
            'repositories_scanned': self.data['repository'].nunique() if not self.data.empty else 0,
            'tools_used': self.data['tool_name'].nunique() if 'tool_name' in self.data.columns else 0
        }
    
    def generate_repository_executive_summary(self) -> pd.DataFrame:
        """
        Generate repository-focused executive summary report.
        
        Returns:
            DataFrame with repository-based executive summary data
        """
        if self.data is None or self.data.empty:
            if self.scoped_repositories:
                print("📊 Generating repository-focused executive summary for scoped repositories...")
                print("⚠️  No alert data found, showing all scoped repositories as clean")
                repo_summary = []
                for repo in self.scoped_repositories:
                    scanned_branch = self.repository_metadata.get(repo, {}).get('default_branch', 'N/A')
                    repo_summary.append({
                        'Repository Name': repo,
                        'Scanned Branch': scanned_branch,
                        'Risk Score': 0,
                        'Total Open': 0,
                        'Critical': 0,
                        'High': 0,
                        'Medium': 0,
                        'Low': 0,
                        'Error': 0,
                        'Warning': 0,
                        'Note': 0,
                        'Total All Alerts': 0,
                        'Fixed Alerts': 0,
                        'Dismissed Alerts': 0
                    })
                summary_df = pd.DataFrame(repo_summary)
                summary_df.insert(0, 'Priority Rank', range(1, len(summary_df) + 1))
                print(f"✅ Repository executive summary generated for {len(summary_df)} repositories (all clean)")
                return summary_df
            return pd.DataFrame()
        
        print("📊 Generating repository-focused executive summary...")
        
        # Filter for open alerts
        open_alerts = self.data[self.data['alert_state'] == 'open'].copy()
        
        print(f"  Debug - Found {len(open_alerts)} open alerts")
        
        # Determine which repositories to include
        if self.scoped_repositories:
            repositories_to_include = self.scoped_repositories
            print(f"  Debug - Including all {len(repositories_to_include)} scoped repositories")
        else:
            if open_alerts.empty:
                print("⚠️  No open alerts found")
                return pd.DataFrame()
            repositories_to_include = open_alerts['repository'].unique().tolist()
            print(f"  Debug - Including {len(repositories_to_include)} repositories with alerts")
        
        # Group by repository and calculate metrics
        repo_summary = []
        
        for repo in repositories_to_include:
            repo_alerts = open_alerts[open_alerts['repository'] == repo] if not open_alerts.empty else pd.DataFrame()
            
            # Get scanned branch
            all_repo_alerts = self.data[self.data['repository'] == repo] if not self.data.empty else pd.DataFrame()
            scanned_branch = 'N/A'
            
            if not all_repo_alerts.empty and 'scanned_branch' in all_repo_alerts.columns:
                branch_values = all_repo_alerts['scanned_branch'].dropna().unique()
                if len(branch_values) > 0:
                    scanned_branch = branch_values[0]
            
            if scanned_branch == 'N/A' and repo in self.repository_metadata:
                scanned_branch = self.repository_metadata[repo].get('default_branch', 'N/A')
            
            if not repo_alerts.empty:
                # Map both severity types - use .copy() to avoid SettingWithCopyWarning
                repo_alerts = repo_alerts.copy()
                severity_col = None
                if 'rule_security_severity_level' in repo_alerts.columns:
                    repo_alerts['combined_severity'] = repo_alerts['rule_security_severity_level'].fillna(
                        repo_alerts['rule_severity']
                    ).str.upper()
                elif 'rule_severity' in repo_alerts.columns:
                    repo_alerts['combined_severity'] = repo_alerts['rule_severity'].str.upper()
                
                severity_counts = repo_alerts['combined_severity'].value_counts() if 'combined_severity' in repo_alerts.columns else {}
                
                risk_score = sum(
                    severity_counts.get(severity, 0) * weight
                    for severity, weight in self.SEVERITY_WEIGHTS.items()
                )
                
                fixed_alerts = all_repo_alerts[all_repo_alerts['alert_state'] == 'fixed']
                dismissed_alerts = all_repo_alerts[all_repo_alerts['alert_state'] == 'dismissed']
                
                repo_summary.append({
                    'Repository Name': repo,
                    'Scanned Branch': scanned_branch,
                    'Risk Score': risk_score,
                    'Total Open': len(repo_alerts),
                    'Critical': severity_counts.get('CRITICAL', 0),
                    'High': severity_counts.get('HIGH', 0),
                    'Medium': severity_counts.get('MEDIUM', 0),
                    'Low': severity_counts.get('LOW', 0),
                    'Error': severity_counts.get('ERROR', 0),
                    'Warning': severity_counts.get('WARNING', 0),
                    'Note': severity_counts.get('NOTE', 0),
                    'Total All Alerts': len(all_repo_alerts),
                    'Fixed Alerts': len(fixed_alerts),
                    'Dismissed Alerts': len(dismissed_alerts)
                })
            else:
                # Repository with no open alerts
                all_fixed = all_repo_alerts[all_repo_alerts['alert_state'] == 'fixed'] if not all_repo_alerts.empty else pd.DataFrame()
                all_dismissed = all_repo_alerts[all_repo_alerts['alert_state'] == 'dismissed'] if not all_repo_alerts.empty else pd.DataFrame()
                
                repo_summary.append({
                    'Repository Name': repo,
                    'Scanned Branch': scanned_branch,
                    'Risk Score': 0,
                    'Total Open': 0,
                    'Critical': 0,
                    'High': 0,
                    'Medium': 0,
                    'Low': 0,
                    'Error': 0,
                    'Warning': 0,
                    'Note': 0,
                    'Total All Alerts': len(all_repo_alerts),
                    'Fixed Alerts': len(all_fixed),
                    'Dismissed Alerts': len(all_dismissed)
                })
        
        # Create DataFrame and sort by risk score
        summary_df = pd.DataFrame(repo_summary)
        summary_df = summary_df.sort_values('Risk Score', ascending=False).reset_index(drop=True)
        
        # Add priority ranking
        summary_df.insert(0, 'Priority Rank', range(1, len(summary_df) + 1))
        
        clean_repos = len(summary_df[summary_df['Total Open'] == 0])
        alert_repos = len(summary_df[summary_df['Total Open'] > 0])
        
        print(f"✅ Repository executive summary generated for {len(summary_df)} repositories")
        if self.scoped_repositories:
            print(f"   📊 Repositories with alerts: {alert_repos}")
            print(f"   ✅ Clean repositories (zero alerts): {clean_repos}")
        
        return summary_df
    
    def generate_detailed_report(self) -> pd.DataFrame:
        """
        Generate detailed technical report with all alert information.
        
        Returns:
            DataFrame with detailed alert data
        """
        if self.data is None or self.data.empty:
            return pd.DataFrame()
        
        print("🔍 Generating detailed technical report...")
        
        # Select and rename columns for the detailed report
        detailed_columns = {
            'repository': 'Repository Name',
            'scanned_branch': 'Scanned Branch',
            'alert_state': 'Status',
            'rule_security_severity_level': 'Security Severity',
            'rule_severity': 'Rule Severity',
            'rule_name': 'Alert Title',
            'rule_id': 'Rule ID',
            'rule_description': 'Description',
            'tool_name': 'Tool',
            'tool_version': 'Tool Version',
            'file_path': 'File Path',
            'start_line': 'Line',
            'created_at': 'Created At',
            'fixed_at': 'Fixed At',
            'dismissed_at': 'Dismissed At',
            'dismissed_by': 'Dismissed By',
            'dismissed_reason': 'Dismissal Reason',
            'dismissed_comment': 'Dismissal Comment',
            'alert_url': 'Alert URL',
            'alert_age_days': 'Age (Days)',
            'rule_tags': 'Tags'
        }
        
        # Create detailed report
        detailed_df = self.data.copy()
        
        # Ensure all required columns exist
        for col in detailed_columns.keys():
            if col not in detailed_df.columns:
                detailed_df[col] = 'N/A'
        
        # Select and rename columns
        detailed_df = detailed_df[list(detailed_columns.keys())].rename(columns=detailed_columns)
        
        # Clean and format data
        detailed_df['Status'] = detailed_df['Status'].str.upper()
        detailed_df['Security Severity'] = detailed_df['Security Severity'].fillna('N/A').str.upper()
        detailed_df['Rule Severity'] = detailed_df['Rule Severity'].fillna('N/A').str.upper()
        
        # Format dates
        date_columns = ['Created At', 'Fixed At', 'Dismissed At']
        for col in date_columns:
            if col in detailed_df.columns:
                detailed_df[col] = detailed_df[col].fillna('N/A')
        
        # Sort by severity
        detailed_df = detailed_df.sort_values('Age (Days)', ascending=False, na_position='last')
        detailed_df = detailed_df.reset_index(drop=True)
        
        print(f"✅ Detailed report generated with {len(detailed_df)} alerts")
        return detailed_df
    
    def save_reports(self, repo_executive_df: pd.DataFrame, detailed_df: pd.DataFrame) -> Optional[str]:
        """
        Save all reports to files.
        
        Args:
            repo_executive_df: Repository executive summary DataFrame
            detailed_df: Detailed alerts DataFrame
            
        Returns:
            Path to reports directory
        """
        try:
            # Create timestamped directory
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if self.active_scope:
                dir_name = f"{self.active_scope}_code_scanning_reports_{timestamp}"
            else:
                dir_name = f"code_scanning_reports_{timestamp}"
            
            reports_dir = Path(self.output_dir) / dir_name
            reports_dir.mkdir(parents=True, exist_ok=True)
            
            print(f"📁 Creating comprehensive reports in: {reports_dir}")
            
            # Save CSVs
            if not repo_executive_df.empty:
                repo_executive_df.to_csv(reports_dir / "executive_summary.csv", index=False)
                print("✅ Executive CSV: executive_summary.csv")
            
            if not detailed_df.empty:
                detailed_df.to_csv(reports_dir / "detailed_alerts.csv", index=False)
                print("✅ Detailed CSV: detailed_alerts.csv")
            
            # Save Excel files with formatting
            try:
                self._save_excel_reports(repo_executive_df, detailed_df, reports_dir)
            except Exception as e:
                print(f"⚠️  Could not create Excel reports: {e}")
            
            # Create README
            self._create_readme(reports_dir, repo_executive_df, detailed_df)
            
            print(f"✅ Comprehensive report suite generated successfully in: {reports_dir}")
            return str(reports_dir)
            
        except Exception as e:
            print(f"❌ Error saving reports: {e}")
            return None
    
    def _save_excel_reports(self, repo_executive_df: pd.DataFrame, detailed_df: pd.DataFrame, output_dir: Path):
        """
        Save formatted Excel reports with styling.
        
        Args:
            repo_executive_df: Repository executive summary DataFrame
            detailed_df: Detailed alerts DataFrame
            output_dir: Output directory path
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Save Executive Summary Excel
        if not repo_executive_df.empty:
            exec_filename = output_dir / "executive_summary.xlsx"
            with pd.ExcelWriter(exec_filename, engine='openpyxl') as writer:
                repo_executive_df.to_excel(writer, sheet_name='Executive Summary', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Executive Summary']
                
                # Define styles
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                # Style header row
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                              min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                
                # Apply severity-based color coding (matching security_report_generator.py)
                try:
                    # Create column mapping
                    col_map = {col: idx+1 for idx, col in enumerate(repo_executive_df.columns)}
                    
                    for row in range(2, len(repo_executive_df)+2):
                        # Critical column - Purple (B1A0C7)
                        if 'Critical' in col_map:
                            val = worksheet.cell(row=row, column=col_map['Critical']).value
                            if val and val != 0:
                                worksheet.cell(row=row, column=col_map['Critical']).fill = PatternFill(
                                    start_color='B1A0C7', end_color='B1A0C7', fill_type='solid')
                        
                        # High column - Red (FF0000)
                        if 'High' in col_map:
                            val = worksheet.cell(row=row, column=col_map['High']).value
                            if val and val != 0:
                                worksheet.cell(row=row, column=col_map['High']).fill = PatternFill(
                                    start_color='FF0000', end_color='FF0000', fill_type='solid')
                                worksheet.cell(row=row, column=col_map['High']).font = Font(color='FFFFFF', bold=True)
                        
                        # Medium column - Orange (F79646)
                        if 'Medium' in col_map:
                            val = worksheet.cell(row=row, column=col_map['Medium']).value
                            if val and val != 0:
                                worksheet.cell(row=row, column=col_map['Medium']).fill = PatternFill(
                                    start_color='F79646', end_color='F79646', fill_type='solid')
                        
                        # Low column - Light Blue (DAEEF3)
                        if 'Low' in col_map:
                            val = worksheet.cell(row=row, column=col_map['Low']).value
                            if val and val != 0:
                                worksheet.cell(row=row, column=col_map['Low']).fill = PatternFill(
                                    start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
                        
                        # Repository Name cell - Green (92D050) if no open alerts, Yellow (FFFF00) if has alerts
                        if 'Total Open' in col_map and 'Repository Name' in col_map:
                            open_val = worksheet.cell(row=row, column=col_map['Total Open']).value
                            repo_cell = worksheet.cell(row=row, column=col_map['Repository Name'])
                            if open_val == 0:
                                repo_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                            elif open_val and open_val > 0:
                                repo_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                
                except Exception as e:
                    print(f"⚠️  Severity color coding error: {e}")
                
                # Freeze top row
                worksheet.freeze_panes = 'A2'
            
            print("✅ Executive Excel: executive_summary.xlsx")
        
        # Save Detailed Alerts Excel
        if not detailed_df.empty:
            detailed_filename = output_dir / "detailed_alerts.xlsx"
            with pd.ExcelWriter(detailed_filename, engine='openpyxl') as writer:
                detailed_df.to_excel(writer, sheet_name='Detailed Alerts', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Detailed Alerts']
                
                # Define styles
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                # Style header row
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                              min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                
                # Apply severity-based color coding for detailed alerts
                try:
                    # Severity color scheme matching security_report_generator.py
                    severity_colors = {
                        'CRITICAL': {'bg': 'B1A0C7', 'font': '000000'},
                        'HIGH': {'bg': 'FF0000', 'font': 'FFFFFF'},
                        'MEDIUM': {'bg': 'F79646', 'font': '000000'},
                        'LOW': {'bg': 'DAEEF3', 'font': '000000'},
                        'ERROR': {'bg': 'FF0000', 'font': 'FFFFFF'},
                        'WARNING': {'bg': 'F79646', 'font': '000000'},
                        'NOTE': {'bg': 'DAEEF3', 'font': '000000'}
                    }
                    
                    # Status color scheme
                    status_colors = {
                        'OPEN': {'bg': 'FFFF00', 'font': '000000'},
                        'FIXED': {'bg': '92D050', 'font': '000000'},
                        'DISMISSED': {'bg': 'D3D3D3', 'font': '000000'}
                    }
                    
                    # Create column mapping
                    col_map = {col: idx+1 for idx, col in enumerate(detailed_df.columns)}
                    
                    for row in range(2, len(detailed_df)+2):
                        # Color code Security Severity column
                        if 'Security Severity' in col_map:
                            severity_cell = worksheet.cell(row=row, column=col_map['Security Severity'])
                            if severity_cell.value:
                                severity = str(severity_cell.value).upper()
                                if severity in severity_colors:
                                    colors = severity_colors[severity]
                                    severity_cell.fill = PatternFill(
                                        start_color=colors['bg'], end_color=colors['bg'], fill_type='solid')
                                    severity_cell.font = Font(color=colors['font'], bold=True)
                        
                        # Color code Rule Severity column
                        if 'Rule Severity' in col_map:
                            rule_severity_cell = worksheet.cell(row=row, column=col_map['Rule Severity'])
                            if rule_severity_cell.value:
                                severity = str(rule_severity_cell.value).upper()
                                if severity in severity_colors:
                                    colors = severity_colors[severity]
                                    rule_severity_cell.fill = PatternFill(
                                        start_color=colors['bg'], end_color=colors['bg'], fill_type='solid')
                                    rule_severity_cell.font = Font(color=colors['font'], bold=True)
                        
                        # Color code Status column
                        if 'Status' in col_map:
                            status_cell = worksheet.cell(row=row, column=col_map['Status'])
                            if status_cell.value:
                                status = str(status_cell.value).upper()
                                if status in status_colors:
                                    colors = status_colors[status]
                                    status_cell.fill = PatternFill(
                                        start_color=colors['bg'], end_color=colors['bg'], fill_type='solid')
                                    status_cell.font = Font(color=colors['font'], bold=True)
                
                except Exception as e:
                    print(f"⚠️  Detailed severity color coding error: {e}")
                
                # Freeze top row
                worksheet.freeze_panes = 'A2'
            
            print("✅ Detailed Excel: detailed_alerts.xlsx")
    
    def _create_readme(self, output_dir: Path, repo_df: pd.DataFrame, detailed_df: pd.DataFrame):
        """Create README file for the reports."""
        readme_content = f"""# Code Scanning Security Report
        
Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Scope: {self.active_scope if self.active_scope else "All Repositories"}

## Summary

- **Total Repositories Scanned:** {self.stats.get('repositories_scanned', 0)}
- **Total Alerts Found:** {self.stats.get('total_alerts', 0)}
- **Open Alerts:** {self.stats.get('open_alerts', 0)}
- **Fixed Alerts:** {self.stats.get('fixed_alerts', 0)}
- **Dismissed Alerts:** {self.stats.get('dismissed_alerts', 0)}

## Report Files

### CSV Files (Plain Text)
1. **executive_summary.csv** - High-level summary by repository
2. **detailed_alerts.csv** - Complete alert details

### Excel Files (Formatted)
1. **executive_summary.xlsx** - Formatted executive summary with styling
2. **detailed_alerts.xlsx** - Formatted detailed alerts with styling

## Severity Breakdown

Open alerts by severity level:
"""
        
        if not detailed_df.empty and 'Security Severity' in detailed_df.columns:
            open_df = detailed_df[detailed_df['Status'] == 'OPEN']
            if not open_df.empty:
                severity_counts = open_df['Security Severity'].value_counts()
                for severity, count in severity_counts.items():
                    readme_content += f"- {severity}: {count}\n"
        
        readme_content += "\n## Top Repositories by Risk\n\n"
        if not repo_df.empty:
            top_repos = repo_df.head(5)
            for _, row in top_repos.iterrows():
                readme_content += f"{row['Priority Rank']}. {row['Repository Name']} (Risk Score: {row['Risk Score']}, Open Alerts: {row['Total Open']})\n"
        
        with open(output_dir / "README.md", 'w') as f:
            f.write(readme_content)
        
        print("✅ README: README.md")


def main():
    """Example usage."""
    generator = CodeScanningReportGenerator()
    
    # Load data
    if generator.load_alert_data("temp_code_scanning_20241111_120000.json"):
        # Generate reports
        repo_summary = generator.generate_repository_executive_summary()
        detailed = generator.generate_detailed_report()
        
        # Save reports
        generator.save_reports(repo_summary, detailed)


if __name__ == "__main__":
    main()
